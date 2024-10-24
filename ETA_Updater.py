from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl
import warnings
import variables
import time


# create a new instance of Chrome WebDriver
chrome_options = Options()
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--disable-gpu")
#chrome_options.add_argument("--headless")
chrome_options.page_load_strategy = ("--eager")
chrome_options.add_argument("--disable-blink-features=AutomationControlled")




#INITIALIZE THE DRIVER FOR MSC, accepting coockies(Needed just the first time) so when the function iterates skips the cookie check
msc_driver = webdriver.Chrome(variables.web_driver_location, options=chrome_options)
msc_driver.get(variables.MSC_website)
cookie = WebDriverWait(msc_driver, 5).until(EC.presence_of_element_located((By.ID, variables.MSC_cookie_button)))
#Acceptcoockies
cookie.click()



#INITIALIZE THE DRIVER FOR CMA, accepting coockies(Needed just the first time) so when the function iterates skips the cookie check

#cookie = WebDriverWait(msc_driver, 5).until(EC.presence_of_element_located((By.ID, variables.MSC_cookie_button)))
#Acceptcoockies
#cookie.click()


#SET UP REPORT FOR EXTRACTION
excel_report = variables.imput_report_location

#Open and parse the exel file and then srape the ETA from MSC website via the dedicated function
def copy_first_cell_values(filename):
    # Open the workbook
    warnings.simplefilter("ignore")
    workbook = openpyxl.load_workbook(filename)
    # Select the first sheet
    sheet = workbook.active
    # Loop through the rows
    for row in sheet.iter_rows(min_row=2, max_col=1):
        # Find the number of containers to check
        n_rows = sheet.max_row - 1
        # Get the value of the first cell in the row(The container number)
        cell_value = row[0].value
        #Run the ETA scrape function
        ETA = MSC_ETA_CHECK(cell_value)
        # Add ETA value to the 5th column of the current row
        last_row_index = row[0].row
        sheet.cell(row=last_row_index, column=5).value = ETA
        remaining = n_rows-last_row_index+2
        print("Need to check "+ str(remaining)+" Containers")
    # Close the workbook
    workbook.save(filename = variables.output_report_location)
    workbook.close()


def MSC_ETA_CHECK(Container_number):
    try:
        # navigate to the web page
        msc_driver.get(variables.MSC_website)
        #Imput the tracking number into the imput field
        input_field = msc_driver.find_element_by_xpath(variables.MSC_cntr_imput)
        input_field.send_keys(Container_number)
        #Click search button
        search_button = msc_driver.find_element_by_xpath(variables.MSC_search_button)
        search_button.click()
        #Capture the ETA, wait for it to become visible
        eta = WebDriverWait(msc_driver, 15).until(
        EC.presence_of_element_located((By.XPATH, variables.MSC_eta_object)))
        eta_value = eta.text
        return eta_value
    except TimeoutException:
        print("Timeout exception occurred. Trying again in 5 seconds...")
        time.sleep(5)
        return MSC_ETA_CHECK(Container_number)



#Run the function that starts the process of checking ETA
copy_first_cell_values(excel_report)

#ADD function to clean the table after scraping the ETA and leave only the rows where our ETA is wrong and needs to be updated
#Add function for CMA and Hapag Lloyd