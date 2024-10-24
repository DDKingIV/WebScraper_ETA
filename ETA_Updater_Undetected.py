from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl
import warnings
import variables
import time
from datetime import datetime

#Setting up the driver with needed options
chrome_options = Options()
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--headless=new")
chrome_options.add_argument("--no-sandbox")


#SET UP REPORT FOR EXTRACTION
excel_report = variables.ALL_CARRIERS_test

#Function to clean up report and save it
def delete_rows_with_equal_dates(file_path):
    # Load the Excel file
    print("Preparing the report for ingestion")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    # Iterate over the rows in reverse order
    for row in reversed(list(ws.iter_rows(min_row=2))):
        # Get the values in columns 3 and 5
        col3_value = row[2].value
        col5_value = row[4].value
        # Check if the value in column 5 is "Unknown"
        if isinstance(col5_value, str) and col5_value.strip().lower() == "unknown,":
            ws.delete_rows(row[0].row)
            continue
        # Convert the values to datetime objects if they're strings
        if isinstance(col3_value, str):
            col3_value = datetime.strptime(col3_value, '%Y-%m-%d')
        if isinstance(col5_value, str):
            col5_value = datetime.strptime(col5_value, '%Y-%m-%d')
        # Compare only the date components of the values
        if col3_value.date() == col5_value.date():
            # Delete the row
            ws.delete_rows(row[0].row)
    # Save the changes
    wb.save(variables.save_path)
    print("All done now!")


def copy_first_cell_values(filename):
    # Open the workbook
    warnings.simplefilter("ignore")
    workbook = openpyxl.load_workbook(filename)
    # Select the first sheet
    sheet = workbook.active
    # Loop through the rows
    for row in sheet.iter_rows(min_row=2, max_col=1):
        # Find the number of containers to check
        n_rows = sheet.max_row - 1
        # Get the value of the first cell in the row(The container number)
        cell_value = row[0].value
        #Run the ETA scrape function
        ETA_long = ALL_CARRIERS_ETA_CHECK(cell_value)
        #extract only the date from the string
        ETA = ETA_long[-10:]
        last_row_index = row[0].row
        #add date to the report
        sheet.cell(row=last_row_index, column=5).value = ETA
        #code to update user on the status of the process
        remaining = n_rows-last_row_index+2
        print("Need to check "+ str(remaining)+" Containers")
    workbook.save(filename=variables.output_report_location)
    workbook.close()
    delete_rows_with_equal_dates(variables.output_report_location)

def ALL_CARRIERS_ETA_CHECK(Container_number):
    try:
        # navigate to the web page
        SR_driver.get(variables.SR_website)
        #Imput the tracking number into the imput field
        input_field = WebDriverWait(SR_driver, 15).until(
        EC.presence_of_element_located((By.XPATH, variables.SR_cntr_imput)))
        input_field.send_keys(Container_number)
        #Click search button
        search_button = SR_driver.find_element_by_xpath(variables.SR_search_button)
        search_button.click()
        #Capture the ETA, wait for it to become visible
        eta = WebDriverWait(SR_driver, 20).until(
        EC.presence_of_element_located((By.XPATH, variables.SR_eta_object)))
        eta_value = eta.text
        return eta_value
    except TimeoutException:
        print("Timeout exception occurred. Trying again in 5 seconds...")
        time.sleep(5)
        return ALL_CARRIERS_ETA_CHECK(Container_number)


#INITIALIZE THE DRIVER, making a login procedure each time the driver is initialized because attributes are not saved between sessions
SR_driver = webdriver.Chrome(variables.web_driver_location,options=chrome_options)
SR_driver.get(variables.SR_website)
log1 = WebDriverWait(SR_driver, 15).until(
        EC.presence_of_element_located((By.XPATH, variables.SR_Menu_btn)))
log1.click()
log2  = WebDriverWait(SR_driver, 15).until(
        EC.presence_of_element_located((By.XPATH, variables.SR_Login_btn)))
log2.click()
email = WebDriverWait(SR_driver, 15).until(
        EC.presence_of_element_located((By.XPATH, variables.SR_email_imput)))
email.send_keys(variables.SR_email)
psw = WebDriverWait(SR_driver, 15).until(
        EC.presence_of_element_located((By.XPATH, variables.SR_psw_Imput)))
psw.send_keys(variables.SR_psw)
conf = WebDriverWait(SR_driver, 15).until(
        EC.presence_of_element_located((By.XPATH, variables.SR_confirm_login)))
conf.click()

#Run the main function that extracts the data
copy_first_cell_values(excel_report)

